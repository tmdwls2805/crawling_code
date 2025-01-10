import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

def post_processing_text(content):
    return re.sub(r'[\r\n\t]', '', content).strip()


def post_processing_apply_content(apply_content):
    form_data = []
    cleaned_list = [item.strip() for item in apply_content.split() if item.strip()]
    original_string = " ".join(cleaned_list)
    split_result = re.split(r'(○ |- )', original_string)
    final_result = []
    temp = ""
    for part in split_result:
        if re.match(r'(○ |- )', part):  # 구분자인 경우
            if temp:  # temp에 기존 값이 있다면 추가
                final_result.append(temp.strip())
            temp = part  # 구분자로 초기화
        else:
            temp += part  # 기존 temp에 누적

    if temp:  # 마지막 부분 추가
        final_result.append(temp.strip())
    for split in final_result:
        if split != '':
            form_data.append(split.strip())
    return form_data


def convert_list_to_multiline(data):
    """
    리스트 데이터를 줄바꿈('\n')으로 연결된 문자열로 변환.
    리스트가 아닌 데이터는 그대로 반환.
    """
    if isinstance(data, list):
        return '\n'.join([str(item) for item in data])  # 리스트를 줄바꿈으로 연결
    return data  # 리스트가 아니면 그대로 반환


def apply_text_wrap_and_adjust(file_path, sheet_names):
    """
    엑셀 파일의 모든 셀에 텍스트 줄 바꿈 속성을 적용하고, 셀 높이를 글씨 크기에 맞게 조정.
    """
    wb = load_workbook(file_path)
    for sheet_name in sheet_names:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value:  # 셀이 비어 있지 않을 때만 줄 바꿈 및 정렬 적용
                    cell.alignment = Alignment(wrap_text=True, vertical='top')  # 줄 바꿈 및 위쪽 정렬
                    cell.font = Font(size=12)  # 글씨 크기 조정 (예: 12pt)

        # 행 높이 조정
        for row in sheet.iter_rows():
            max_line_count = max(
                len(str(cell.value).split('\n')) if cell.value else 1 for cell in row
            )
            sheet.row_dimensions[row[0].row].height = max_line_count * 15  # 줄 수에 따라 높이 조정

    wb.save(file_path)
